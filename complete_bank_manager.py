#!/usr/bin/env python3
"""
Complete Bank Manager - All Features Included
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class CompleteDatabase:
    """Complete database with all features"""
    
    def __init__(self, master_password):
        self.master_password = master_password
        self.db_path = 'complete_bank_manager.db'
        self._init_database()
    
    def _init_database(self):
        """Initialize database with all fields"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bank_cards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_name TEXT NOT NULL,
                branch_name TEXT NOT NULL,
                ifsc_code TEXT NOT NULL,
                account_number TEXT NOT NULL,
                atm_number TEXT NOT NULL,
                pin TEXT NOT NULL,
                validity_start TEXT NOT NULL,
                validity_end TEXT NOT NULL,
                cvv TEXT NOT NULL,
                card_type TEXT NOT NULL,
                card_network TEXT NOT NULL,
                family_member TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def mask_card_number(self, card_number):
        """Mask card number"""
        if len(card_number) >= 16:
            return f"{card_number[:4]} **** **** {card_number[-4:]}"
        return "**** **** **** ****"
    
    def mask_cvv(self, cvv):
        """Mask CVV"""
        if len(cvv) == 3:
            return "***"
        return "**"
    
    def add_card(self, bank_name, branch_name, ifsc_code, account_number, atm_number, pin, 
                 validity_start, validity_end, cvv, card_type, card_network, family_member):
        """Add new bank card"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        now = datetime.now().isoformat()
        
        cursor.execute('''
            INSERT INTO bank_cards 
            (bank_name, branch_name, ifsc_code, account_number, atm_number, pin, 
             validity_start, validity_end, cvv, card_type, card_network, 
             family_member, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            bank_name, branch_name, ifsc_code, account_number, atm_number, pin, 
            validity_start, validity_end, cvv, card_type, card_network, 
            family_member, now, now
        ))
        
        card_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return card_id
    
    def get_all_cards(self):
        """Get all cards with masked sensitive data"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, bank_name, branch_name, ifsc_code, account_number, atm_number, 
                   pin, validity_start, validity_end, cvv, card_type, card_network,
                   family_member, created_at, updated_at
            FROM bank_cards
            ORDER BY created_at DESC
        ''')
        
        cards = []
        for row in cursor.fetchall():
            card_data = {
                'id': row[0],
                'bank_name': row[1],
                'branch_name': row[2],
                'ifsc_code': row[3],
                'account_number': self.mask_card_number(row[4]),
                'atm_number': self.mask_card_number(row[5]),
                'pin': self.mask_cvv(row[6]),
                'validity_start': row[7],
                'validity_end': row[8],
                'cvv': self.mask_cvv(row[9]),
                'card_type': row[10],
                'card_network': row[11],
                'family_member': row[12],
                'created_at': row[13],
                'updated_at': row[14]
            }
            cards.append(card_data)
        
        conn.close()
        return cards
    
    def get_card_by_id(self, card_id):
        """Get card by ID for editing"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, bank_name, branch_name, ifsc_code, account_number, atm_number, 
                   pin, validity_start, validity_end, cvv, card_type, card_network,
                   family_member, created_at, updated_at
            FROM bank_cards WHERE id = ?
        ''', (card_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return {
                'id': row[0],
                'bank_name': row[1],
                'branch_name': row[2],
                'ifsc_code': row[3],
                'account_number': row[4],
                'atm_number': row[5],
                'pin': row[6],
                'validity_start': row[7],
                'validity_end': row[8],
                'cvv': row[9],
                'card_type': row[10],
                'card_network': row[11],
                'family_member': row[12],
                'created_at': row[13],
                'updated_at': row[14]
            }
        return None
    
    def update_card(self, card_id, bank_name, branch_name, ifsc_code, account_number, 
                   atm_number, pin, validity_start, validity_end, cvv, card_type, 
                   card_network, family_member):
        """Update existing card"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        now = datetime.now().isoformat()
        
        cursor.execute('''
            UPDATE bank_cards 
            SET bank_name=?, branch_name=?, ifsc_code=?, account_number=?, atm_number=?, 
                pin=?, validity_start=?, validity_end=?, cvv=?, card_type=?, 
                card_network=?, family_member=?, updated_at=?
            WHERE id=?
        ''', (
            bank_name, branch_name, ifsc_code, account_number, atm_number, pin,
            validity_start, validity_end, cvv, card_type, card_network, 
            family_member, now, card_id
        ))
        
        conn.commit()
        conn.close()
    
    def delete_card(self, card_id):
        """Delete card by ID"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM bank_cards WHERE id = ?', (card_id,))
        conn.commit()
        conn.close()
    
    def get_all_cards_unmasked(self):
        """Get all cards with unmasked sensitive data for export"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, bank_name, branch_name, ifsc_code, account_number, atm_number, 
                   pin, validity_start, validity_end, cvv, card_type, card_network,
                   family_member, created_at, updated_at
            FROM bank_cards
            ORDER BY created_at DESC
        ''')
        
        cards = []
        for row in cursor.fetchall():
            card_data = {
                'id': row[0],
                'bank_name': row[1],
                'branch_name': row[2],
                'ifsc_code': row[3],
                'account_number': row[4],  # Unmasked
                'atm_number': row[5],      # Unmasked
                'pin': row[6],             # Unmasked
                'validity_start': row[7],
                'validity_end': row[8],
                'cvv': row[9],             # Unmasked
                'card_type': row[10],
                'card_network': row[11],
                'family_member': row[12],
                'created_at': row[13],
                'updated_at': row[14]
            }
            cards.append(card_data)
        
        conn.close()
        return cards

    def export_to_excel(self, filename):
        """Export cards to Excel file"""
        cards = self.get_all_cards_unmasked()
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Cards"
        
        # Define headers
        headers = [
            "ID", "Bank Name", "Branch Name", "IFSC Code", "Account Number", 
            "ATM Number", "PIN", "Valid From", "Valid Until", "CVV", 
            "Card Type", "Card Network", "Family Member", "Created At", "Updated At"
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row, card in enumerate(cards, 2):
            ws.cell(row=row, column=1, value=card['id'])
            ws.cell(row=row, column=2, value=card['bank_name'])
            ws.cell(row=row, column=3, value=card['branch_name'])
            ws.cell(row=row, column=4, value=card['ifsc_code'])
            ws.cell(row=row, column=5, value=card['account_number'])
            ws.cell(row=row, column=6, value=card['atm_number'])
            ws.cell(row=row, column=7, value=card['pin'])
            ws.cell(row=row, column=8, value=card['validity_start'])
            ws.cell(row=row, column=9, value=card['validity_end'])
            ws.cell(row=row, column=10, value=card['cvv'])
            ws.cell(row=row, column=11, value=card['card_type'])
            ws.cell(row=row, column=12, value=card['card_network'])
            ws.cell(row=row, column=13, value=card['family_member'])
            ws.cell(row=row, column=14, value=card['created_at'])
            ws.cell(row=row, column=15, value=card['updated_at'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws['A1'] = "Bank Cards Export Summary"
        summary_ws['A1'].font = Font(bold=True, size=14)
        
        summary_ws['A3'] = "Export Date:"
        summary_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        summary_ws['A4'] = "Total Cards:"
        summary_ws['B4'] = len(cards)
        
        summary_ws['A6'] = "Card Types:"
        card_types = {}
        for card in cards:
            card_type = card['card_type']
            card_types[card_type] = card_types.get(card_type, 0) + 1
        
        row = 7
        for card_type, count in card_types.items():
            summary_ws[f'A{row}'] = f"{card_type}:"
            summary_ws[f'B{row}'] = count
            row += 1
        
        # Save the workbook
        wb.save(filename)
        return len(cards)
    
    def import_from_excel(self, filename):
        """Import cards from Excel file"""
        try:
            wb = load_workbook(filename, data_only=True)
            ws = wb.active
            
            imported_count = 0
            skipped_count = 0
            
            # Skip header row, start from row 2
            for row in range(2, ws.max_row + 1):
                try:
                    # Read data from Excel
                    bank_name = str(ws.cell(row=row, column=2).value or "").strip()
                    branch_name = str(ws.cell(row=row, column=3).value or "").strip()
                    ifsc_code = str(ws.cell(row=row, column=4).value or "").strip()
                    account_number = str(ws.cell(row=row, column=5).value or "").strip()
                    atm_number = str(ws.cell(row=row, column=6).value or "").strip()
                    pin = str(ws.cell(row=row, column=7).value or "").strip()
                    validity_start = str(ws.cell(row=row, column=8).value or "").strip()
                    validity_end = str(ws.cell(row=row, column=9).value or "").strip()
                    cvv = str(ws.cell(row=row, column=10).value or "").strip()
                    card_type = str(ws.cell(row=row, column=11).value or "Debit").strip()
                    card_network = str(ws.cell(row=row, column=12).value or "RuPay").strip()
                    family_member = str(ws.cell(row=row, column=13).value or "").strip()
                    
                    # Validate required fields
                    if not all([bank_name, branch_name, ifsc_code, account_number, atm_number, pin, cvv, validity_start, validity_end, family_member]):
                        skipped_count += 1
                        continue
                    
                    # Clean account and ATM numbers (remove spaces)
                    account_number = ''.join(filter(str.isdigit, account_number))
                    atm_number = ''.join(filter(str.isdigit, atm_number))
                    
                    # Validate ATM number length
                    if len(atm_number) != 16:
                        skipped_count += 1
                        continue
                    
                    # Validate date format
                    try:
                        datetime.strptime(validity_start, "%Y-%m-%d")
                        datetime.strptime(validity_end, "%Y-%m-%d")
                    except ValueError:
                        skipped_count += 1
                        continue
                    
                    # Add card to database
                    self.add_card(
                        bank_name, branch_name, ifsc_code, account_number, atm_number, pin,
                        validity_start, validity_end, cvv, card_type, card_network, family_member
                    )
                    imported_count += 1
                    
                except Exception as e:
                    skipped_count += 1
                    continue
            
            return imported_count, skipped_count
            
        except Exception as e:
            raise Exception(f"Failed to import Excel file: {e}")
    
    def export_to_json(self, filename):
        """Export cards to JSON file (legacy support)"""
        cards = self.get_all_cards_unmasked()
        
        export_data = {
            'export_date': datetime.now().isoformat(),
            'version': '1.0.0',
            'total_cards': len(cards),
            'cards': cards
        }
        
        with open(filename, 'w') as f:
            json.dump(export_data, f, indent=2)
        
        return len(cards)

class LoginWindow:
    """Login window"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Complete Bank Manager - Login")
        self.root.geometry("400x200")
        self.root.resizable(False, False)
        
        # Center window
        self.center_window(self.root)
        
        self.database = None
        self.create_widgets()
    
    def center_window(self, window):
        """Center window on screen"""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Create login widgets"""
        # Title
        title = tk.Label(self.root, text="Complete Bank Manager", font=("Arial", 16, "bold"))
        title.pack(pady=20)
        
        # Password frame
        password_frame = ttk.Frame(self.root)
        password_frame.pack(pady=20)
        
        ttk.Label(password_frame, text="Master Password:").pack()
        self.password_entry = ttk.Entry(password_frame, show="*", width=30)
        self.password_entry.pack(pady=5)
        self.password_entry.focus()
        
        # Login button
        login_btn = ttk.Button(password_frame, text="Login", command=self.login)
        login_btn.pack(pady=10)
        
        # Bind Enter key
        self.password_entry.bind('<Return>', lambda e: self.login())
    
    def login(self):
        """Handle login"""
        password = self.password_entry.get()
        
        if len(password) < 8:
            messagebox.showerror("Error", "Password must be at least 8 characters long!")
            return
        
        try:
            self.database = CompleteDatabase(password)
            self.root.destroy()
            self.show_main_window()
        except Exception as e:
            messagebox.showerror("Error", f"Login failed: {e}")
    
    def show_main_window(self):
        """Show main window after login"""
        MainWindow(self.database)
    
    def run(self):
        """Run the login window"""
        self.root.mainloop()

class MainWindow:
    """Main application window"""
    
    def __init__(self, database):
        self.database = database
        self.root = tk.Tk()
        self.root.title("Complete Bank Manager")
        self.root.geometry("1000x700")
        
        # Center window
        self.center_window(self.root)
        
        self.create_widgets()
        self.refresh_cards()
    
    def center_window(self, window):
        """Center window on screen"""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Create main window widgets"""
        # Title
        title = tk.Label(self.root, text="Complete Bank Manager", font=("Arial", 16, "bold"))
        title.pack(pady=10)
        
        # Buttons frame
        button_frame = ttk.Frame(self.root)
        button_frame.pack(pady=10)
        
        add_btn = ttk.Button(button_frame, text="Add Card", command=self.add_card)
        add_btn.pack(side="left", padx=5)
        
        edit_btn = ttk.Button(button_frame, text="Edit Card", command=self.edit_card)
        edit_btn.pack(side="left", padx=5)
        
        refresh_btn = ttk.Button(button_frame, text="Refresh", command=self.refresh_cards)
        refresh_btn.pack(side="left", padx=5)
        
        export_btn = ttk.Button(button_frame, text="Export Excel", command=self.export_cards)
        export_btn.pack(side="left", padx=5)
        
        import_btn = ttk.Button(button_frame, text="Import Excel", command=self.import_cards)
        import_btn.pack(side="left", padx=5)
        
        logout_btn = ttk.Button(button_frame, text="Logout", command=self.logout)
        logout_btn.pack(side="left", padx=5)
        
        # Cards frame
        cards_frame = ttk.Frame(self.root)
        cards_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Treeview for cards
        columns = ("Bank", "Branch", "Account", "ATM", "Type", "Network", "Member", "Valid Until")
        self.tree = ttk.Treeview(cards_frame, columns=columns, show="headings", height=15)
        
        # Configure columns
        self.tree.heading("Bank", text="Bank Name")
        self.tree.heading("Branch", text="Branch")
        self.tree.heading("Account", text="Account Number")
        self.tree.heading("ATM", text="ATM Number")
        self.tree.heading("Type", text="Type")
        self.tree.heading("Network", text="Network")
        self.tree.heading("Member", text="Family Member")
        self.tree.heading("Valid Until", text="Valid Until")
        
        self.tree.column("Bank", width=100)
        self.tree.column("Branch", width=100)
        self.tree.column("Account", width=120)
        self.tree.column("ATM", width=120)
        self.tree.column("Type", width=50)
        self.tree.column("Network", width=70)
        self.tree.column("Member", width=80)
        self.tree.column("Valid Until", width=80)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(cards_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind double-click for editing
        self.tree.bind("<Double-1>", self.edit_card)
        
        # Bind right-click for context menu
        self.tree.bind("<Button-3>", self.show_context_menu)
        
        # Status bar
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill="x", padx=20, pady=5)
        
        self.status_var = tk.StringVar()
        self.status_var.set("Ready - Select a card to edit or right-click for options")
        status_label = ttk.Label(status_frame, textvariable=self.status_var, relief="sunken")
        status_label.pack(fill="x")
        
        # Bind selection change to update status
        self.tree.bind("<<TreeviewSelect>>", self.update_status)
    
    def refresh_cards(self):
        """Refresh the cards display"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Get cards from database
        cards = self.database.get_all_cards()
        
        # Add cards to treeview
        for card in cards:
            self.tree.insert("", "end", values=(
                card['bank_name'],
                card['branch_name'],
                card['account_number'],
                card['atm_number'],
                card['card_type'],
                card['card_network'],
                card['family_member'],
                card['validity_end']
            ), tags=(card['id'],))
    
    def add_card(self):
        """Show add card dialog"""
        AddCardDialog(self.root, self.database, self.refresh_cards)
    
    def edit_card(self, event=None):
        """Edit selected card"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a card to edit!")
            return
        
        item = selection[0]
        card_id = self.tree.item(item, "tags")[0]
        
        # Get card details
        card = self.database.get_card_by_id(int(card_id))
        if card:
            EditCardDialog(self.root, self.database, card, self.refresh_cards)
        else:
            messagebox.showerror("Error", "Card not found!")
    
    def show_context_menu(self, event):
        """Show context menu on right-click"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu = tk.Menu(self.root, tearoff=0)
            self.context_menu.add_command(label="✏️ Edit Card", command=self.edit_card)
            self.context_menu.add_command(label="🗑️ Delete Card", command=self.delete_card)
            self.context_menu.add_separator()
            self.context_menu.add_command(label="📋 Copy Details", command=self.copy_card_details)
            self.context_menu.tk_popup(event.x_root, event.y_root)
    
    def delete_card(self):
        """Delete selected card"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a card to delete!")
            return
        
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this card?"):
            item = selection[0]
            card_id = self.tree.item(item, "tags")[0]
            
            try:
                self.database.delete_card(int(card_id))
                self.refresh_cards()
                messagebox.showinfo("Success", "Card deleted successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete card: {e}")
    
    def copy_card_details(self):
        """Copy card details to clipboard"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a card to copy!")
            return
        
        item = selection[0]
        card_id = self.tree.item(item, "tags")[0]
        
        # Get card details
        card = self.database.get_card_by_id(int(card_id))
        if card:
            details = f"""Bank: {card['bank_name']}
Branch: {card['branch_name']}
IFSC: {card['ifsc_code']}
Account: {card['account_number']}
ATM: {card['atm_number']}
Type: {card['card_type']}
Network: {card['card_network']}
Member: {card['family_member']}
Valid From: {card['validity_start']}
Valid Until: {card['validity_end']}"""
            
            self.root.clipboard_clear()
            self.root.clipboard_append(details)
            messagebox.showinfo("Copied", "Card details copied to clipboard!")
    
    def export_cards(self):
        """Export cards to Excel"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                count = self.database.export_to_excel(filename)
                messagebox.showinfo("Success", f"Exported {count} cards to {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed: {e}")
    
    def import_cards(self):
        """Import cards from Excel"""
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                imported, skipped = self.database.import_from_excel(filename)
                messagebox.showinfo("Import Complete", 
                    f"Successfully imported {imported} cards.\n"
                    f"Skipped {skipped} invalid entries.")
                self.refresh_cards()  # Refresh the display
            except Exception as e:
                messagebox.showerror("Error", f"Import failed: {e}")
    
    def logout(self):
        """Logout and close application"""
        if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            self.root.destroy()
    
    def run(self):
        """Run the main window"""
        self.root.mainloop()

    def update_status(self, event):
        """Update the status bar based on selected card"""
        selection = self.tree.selection()
        if not selection:
            self.status_var.set("Ready - Select a card to edit or right-click for options")
            return
        
        item = selection[0]
        card_id = self.tree.item(item, "tags")[0]
        
        card = self.database.get_card_by_id(int(card_id))
        if card:
            self.status_var.set(f"Selected Card: {card['bank_name']} - {card['account_number']} (ID: {card['id']})")
        else:
            self.status_var.set("Ready - Select a card to edit or right-click for options")

class AddCardDialog:
    """Dialog for adding cards"""
    
    def __init__(self, parent, database, callback):
        self.parent = parent
        self.database = database
        self.callback = callback
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Add New Card")
        self.dialog.geometry("500x700")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.center_window(self.dialog)
        
        self.create_widgets()
    
    def center_window(self, window):
        """Center window on screen"""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Create form widgets"""
        # Title
        title = tk.Label(self.dialog, text="Add New Card", font=("Arial", 16, "bold"))
        title.pack(pady=10)
        
        # Form frame
        form = ttk.Frame(self.dialog)
        form.pack(pady=10, padx=20, fill="both", expand=True)
        
        # Bank Name
        ttk.Label(form, text="Bank Name:").grid(row=0, column=0, sticky="w", pady=5)
        self.bank_name = ttk.Entry(form, width=30)
        self.bank_name.grid(row=0, column=1, pady=5, padx=5)
        
        # Branch Name
        ttk.Label(form, text="Branch Name:").grid(row=1, column=0, sticky="w", pady=5)
        self.branch_name = ttk.Entry(form, width=30)
        self.branch_name.grid(row=1, column=1, pady=5, padx=5)
        self.branch_name.insert(0, "Main Branch")
        
        # IFSC Code
        ttk.Label(form, text="IFSC Code:").grid(row=2, column=0, sticky="w", pady=5)
        self.ifsc_code = ttk.Entry(form, width=30)
        self.ifsc_code.grid(row=2, column=1, pady=5, padx=5)
        
        # Account Number
        ttk.Label(form, text="Account Number:").grid(row=3, column=0, sticky="w", pady=5)
        self.account_number = ttk.Entry(form, width=30)
        self.account_number.grid(row=3, column=1, pady=5, padx=5)
        self.account_number.bind('<KeyRelease>', self.format_account_number)
        
        # ATM Number
        ttk.Label(form, text="ATM Number:").grid(row=4, column=0, sticky="w", pady=5)
        self.atm_number = ttk.Entry(form, width=30)
        self.atm_number.grid(row=4, column=1, pady=5, padx=5)
        self.atm_number.bind('<KeyRelease>', self.format_atm_number)
        
        # PIN
        ttk.Label(form, text="PIN:").grid(row=5, column=0, sticky="w", pady=5)
        self.pin = ttk.Entry(form, width=30, show="*")
        self.pin.grid(row=5, column=1, pady=5, padx=5)
        
        # CVV
        ttk.Label(form, text="CVV:").grid(row=6, column=0, sticky="w", pady=5)
        self.cvv = ttk.Entry(form, width=30, show="*")
        self.cvv.grid(row=6, column=1, pady=5, padx=5)
        
        # Card Type
        ttk.Label(form, text="Card Type:").grid(row=7, column=0, sticky="w", pady=5)
        self.card_type_var = tk.StringVar(value="Debit")
        card_type = ttk.Combobox(form, textvariable=self.card_type_var, 
                                values=["Debit", "Credit"], state="readonly", width=27)
        card_type.grid(row=7, column=1, pady=5, padx=5)
        
        # Card Network
        ttk.Label(form, text="Card Network:").grid(row=8, column=0, sticky="w", pady=5)
        self.card_network_var = tk.StringVar(value="RuPay")
        card_network = ttk.Combobox(form, textvariable=self.card_network_var, 
                                   values=["RuPay", "Visa", "Mastercard", "American Express", 
                                          "Discover", "JCB", "UnionPay", "Diners Club"], 
                                   state="readonly", width=27)
        card_network.grid(row=8, column=1, pady=5, padx=5)
        
        # Validity Start
        ttk.Label(form, text="Valid From:").grid(row=9, column=0, sticky="w", pady=5)
        self.validity_start = ttk.Entry(form, width=30)
        self.validity_start.grid(row=9, column=1, pady=5, padx=5)
        self.validity_start.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # Validity End
        ttk.Label(form, text="Valid Until:").grid(row=10, column=0, sticky="w", pady=5)
        self.validity_end = ttk.Entry(form, width=30)
        self.validity_end.grid(row=10, column=1, pady=5, padx=5)
        self.validity_end.insert(0, (datetime.now() + timedelta(days=365*3)).strftime("%Y-%m-%d"))
        
        # Family Member
        ttk.Label(form, text="Family Member:").grid(row=11, column=0, sticky="w", pady=5)
        self.family_member = ttk.Entry(form, width=30)
        self.family_member.grid(row=11, column=1, pady=5, padx=5)
        
        # Buttons
        buttons = ttk.Frame(self.dialog)
        buttons.pack(pady=20)
        
        save_btn = ttk.Button(buttons, text="Save", command=self.save_card)
        save_btn.pack(side="left", padx=5)
        
        cancel_btn = ttk.Button(buttons, text="Cancel", command=self.dialog.destroy)
        cancel_btn.pack(side="left", padx=5)
    
    def format_atm_number(self, event=None):
        """Format ATM number with spaces"""
        value = self.atm_number.get()
        digits_only = ''.join(filter(str.isdigit, value))
        
        if len(digits_only) > 16:
            digits_only = digits_only[:16]
        
        formatted = ''
        for i in range(0, len(digits_only), 4):
            if formatted:
                formatted += ' '
            formatted += digits_only[i:i+4]
        
        if formatted != value:
            self.atm_number.delete(0, tk.END)
            self.atm_number.insert(0, formatted)
    
    def format_account_number(self, event=None):
        """Format account number with spaces"""
        value = self.account_number.get()
        digits_only = ''.join(filter(str.isdigit, value))
        
        if len(digits_only) > 20:
            digits_only = digits_only[:20]
        
        formatted = ''
        for i in range(0, len(digits_only), 4):
            if formatted:
                formatted += ' '
            formatted += digits_only[i:i+4]
        
        if formatted != value:
            self.account_number.delete(0, tk.END)
            self.account_number.insert(0, formatted)
    
    def save_card(self):
        """Save the new card"""
        # Get values
        bank_name = self.bank_name.get().strip()
        branch_name = self.branch_name.get().strip()
        ifsc_code = self.ifsc_code.get().strip()
        account_number = ''.join(filter(str.isdigit, self.account_number.get().strip()))
        atm_number = ''.join(filter(str.isdigit, self.atm_number.get().strip()))
        pin = self.pin.get()
        cvv = self.cvv.get()
        card_type = self.card_type_var.get()
        card_network = self.card_network_var.get()
        validity_start = self.validity_start.get().strip()
        validity_end = self.validity_end.get().strip()
        family_member = self.family_member.get().strip()
        
        # Validate ATM number length
        if len(atm_number) != 16:
            messagebox.showerror("Error", "ATM number must be exactly 16 digits!")
            return
        
        # Validate
        if not all([bank_name, branch_name, ifsc_code, account_number, atm_number, pin, cvv, validity_start, validity_end, family_member]):
            messagebox.showerror("Error", "All fields are required!")
            return
        
        # Validate date format
        try:
            datetime.strptime(validity_start, "%Y-%m-%d")
            datetime.strptime(validity_end, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Please enter valid dates in YYYY-MM-DD format!")
            return
        
        try:
            card_id = self.database.add_card(
                bank_name, branch_name, ifsc_code, account_number, atm_number, pin,
                validity_start, validity_end, cvv, card_type, card_network, family_member
            )
            messagebox.showinfo("Success", f"Card added successfully! ID: {card_id}")
            self.dialog.destroy()
            self.callback()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add card: {e}")

class EditCardDialog:
    """Dialog for editing cards"""
    
    def __init__(self, parent, database, card, callback):
        self.parent = parent
        self.database = database
        self.card = card
        self.callback = callback
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Edit Card")
        self.dialog.geometry("500x700")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.center_window(self.dialog)
        
        self.create_widgets()
    
    def center_window(self, window):
        """Center window on screen"""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Create form widgets"""
        # Title
        title = tk.Label(self.dialog, text="Edit Card", font=("Arial", 16, "bold"))
        title.pack(pady=10)
        
        # Form frame
        form = ttk.Frame(self.dialog)
        form.pack(pady=10, padx=20, fill="both", expand=True)
        
        # Bank Name
        ttk.Label(form, text="Bank Name:").grid(row=0, column=0, sticky="w", pady=5)
        self.bank_name = ttk.Entry(form, width=30)
        self.bank_name.grid(row=0, column=1, pady=5, padx=5)
        self.bank_name.insert(0, self.card['bank_name'])
        
        # Branch Name
        ttk.Label(form, text="Branch Name:").grid(row=1, column=0, sticky="w", pady=5)
        self.branch_name = ttk.Entry(form, width=30)
        self.branch_name.grid(row=1, column=1, pady=5, padx=5)
        self.branch_name.insert(0, self.card['branch_name'])
        
        # IFSC Code
        ttk.Label(form, text="IFSC Code:").grid(row=2, column=0, sticky="w", pady=5)
        self.ifsc_code = ttk.Entry(form, width=30)
        self.ifsc_code.grid(row=2, column=1, pady=5, padx=5)
        self.ifsc_code.insert(0, self.card['ifsc_code'])
        
        # Account Number
        ttk.Label(form, text="Account Number:").grid(row=3, column=0, sticky="w", pady=5)
        self.account_number = ttk.Entry(form, width=30)
        self.account_number.grid(row=3, column=1, pady=5, padx=5)
        # Format account number for display
        if self.card['account_number'] and len(self.card['account_number']) >= 12:
            formatted_account = ''
            for i in range(0, len(self.card['account_number']), 4):
                if formatted_account:
                    formatted_account += ' '
                formatted_account += self.card['account_number'][i:i+4]
            self.account_number.insert(0, formatted_account)
        else:
            self.account_number.insert(0, self.card['account_number'])
        self.account_number.bind('<KeyRelease>', self.format_account_number)
        
        # ATM Number
        ttk.Label(form, text="ATM Number:").grid(row=4, column=0, sticky="w", pady=5)
        self.atm_number = ttk.Entry(form, width=30)
        self.atm_number.grid(row=4, column=1, pady=5, padx=5)
        # Format ATM number for display
        if self.card['atm_number'] and len(self.card['atm_number']) == 16:
            formatted_atm = f"{self.card['atm_number'][:4]} {self.card['atm_number'][4:8]} {self.card['atm_number'][8:12]} {self.card['atm_number'][12:16]}"
            self.atm_number.insert(0, formatted_atm)
        else:
            self.atm_number.insert(0, self.card['atm_number'])
        self.atm_number.bind('<KeyRelease>', self.format_atm_number)
        
        # PIN
        ttk.Label(form, text="PIN:").grid(row=5, column=0, sticky="w", pady=5)
        self.pin = ttk.Entry(form, width=30, show="*")
        self.pin.grid(row=5, column=1, pady=5, padx=5)
        self.pin.insert(0, self.card['pin'])
        
        # CVV
        ttk.Label(form, text="CVV:").grid(row=6, column=0, sticky="w", pady=5)
        self.cvv = ttk.Entry(form, width=30, show="*")
        self.cvv.grid(row=6, column=1, pady=5, padx=5)
        self.cvv.insert(0, self.card['cvv'])
        
        # Card Type
        ttk.Label(form, text="Card Type:").grid(row=7, column=0, sticky="w", pady=5)
        self.card_type_var = tk.StringVar(value=self.card['card_type'])
        card_type = ttk.Combobox(form, textvariable=self.card_type_var, 
                                values=["Debit", "Credit"], state="readonly", width=27)
        card_type.grid(row=7, column=1, pady=5, padx=5)
        
        # Card Network
        ttk.Label(form, text="Card Network:").grid(row=8, column=0, sticky="w", pady=5)
        self.card_network_var = tk.StringVar(value=self.card['card_network'])
        card_network = ttk.Combobox(form, textvariable=self.card_network_var, 
                                   values=["RuPay", "Visa", "Mastercard", "American Express", 
                                          "Discover", "JCB", "UnionPay", "Diners Club"], 
                                   state="readonly", width=27)
        card_network.grid(row=8, column=1, pady=5, padx=5)
        
        # Validity Start
        ttk.Label(form, text="Valid From:").grid(row=9, column=0, sticky="w", pady=5)
        self.validity_start = ttk.Entry(form, width=30)
        self.validity_start.grid(row=9, column=1, pady=5, padx=5)
        self.validity_start.insert(0, self.card['validity_start'])
        
        # Validity End
        ttk.Label(form, text="Valid Until:").grid(row=10, column=0, sticky="w", pady=5)
        self.validity_end = ttk.Entry(form, width=30)
        self.validity_end.grid(row=10, column=1, pady=5, padx=5)
        self.validity_end.insert(0, self.card['validity_end'])
        
        # Family Member
        ttk.Label(form, text="Family Member:").grid(row=11, column=0, sticky="w", pady=5)
        self.family_member = ttk.Entry(form, width=30)
        self.family_member.grid(row=11, column=1, pady=5, padx=5)
        self.family_member.insert(0, self.card['family_member'])
        
        # Buttons
        buttons = ttk.Frame(self.dialog)
        buttons.pack(pady=20)
        
        save_btn = ttk.Button(buttons, text="Update", command=self.update_card)
        save_btn.pack(side="left", padx=5)
        
        cancel_btn = ttk.Button(buttons, text="Cancel", command=self.dialog.destroy)
        cancel_btn.pack(side="left", padx=5)
    
    def format_atm_number(self, event=None):
        """Format ATM number with spaces"""
        value = self.atm_number.get()
        digits_only = ''.join(filter(str.isdigit, value))
        
        if len(digits_only) > 16:
            digits_only = digits_only[:16]
        
        formatted = ''
        for i in range(0, len(digits_only), 4):
            if formatted:
                formatted += ' '
            formatted += digits_only[i:i+4]
        
        if formatted != value:
            self.atm_number.delete(0, tk.END)
            self.atm_number.insert(0, formatted)
    
    def format_account_number(self, event=None):
        """Format account number with spaces"""
        value = self.account_number.get()
        digits_only = ''.join(filter(str.isdigit, value))
        
        if len(digits_only) > 20:
            digits_only = digits_only[:20]
        
        formatted = ''
        for i in range(0, len(digits_only), 4):
            if formatted:
                formatted += ' '
            formatted += digits_only[i:i+4]
        
        if formatted != value:
            self.account_number.delete(0, tk.END)
            self.account_number.insert(0, formatted)
    
    def update_card(self):
        """Update the card"""
        # Get values
        bank_name = self.bank_name.get().strip()
        branch_name = self.branch_name.get().strip()
        ifsc_code = self.ifsc_code.get().strip()
        account_number = ''.join(filter(str.isdigit, self.account_number.get().strip()))
        atm_number = ''.join(filter(str.isdigit, self.atm_number.get().strip()))
        pin = self.pin.get()
        cvv = self.cvv.get()
        card_type = self.card_type_var.get()
        card_network = self.card_network_var.get()
        validity_start = self.validity_start.get().strip()
        validity_end = self.validity_end.get().strip()
        family_member = self.family_member.get().strip()
        
        # Validate ATM number length
        if len(atm_number) != 16:
            messagebox.showerror("Error", "ATM number must be exactly 16 digits!")
            return
        
        # Validate
        if not all([bank_name, branch_name, ifsc_code, account_number, atm_number, pin, cvv, validity_start, validity_end, family_member]):
            messagebox.showerror("Error", "All fields are required!")
            return
        
        # Validate date format
        try:
            datetime.strptime(validity_start, "%Y-%m-%d")
            datetime.strptime(validity_end, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Please enter valid dates in YYYY-MM-DD format!")
            return
        
        try:
            self.database.update_card(
                self.card['id'], bank_name, branch_name, ifsc_code, account_number, 
                atm_number, pin, validity_start, validity_end, cvv, card_type, 
                card_network, family_member
            )
            messagebox.showinfo("Success", "Card updated successfully!")
            self.dialog.destroy()
            self.callback()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update card: {e}")

def main():
    """Main function"""
    app = LoginWindow()
    app.run()

if __name__ == '__main__':
    main() 